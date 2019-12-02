

# excel file spreadsheet automation

# import openpyxl as xl

# wb = xl.load_workbook("transactions.xlsx")
# sheet = wb["Sheet1"]
# cell = sheet["a1"]
# # cell = sheet.cell(1, 1)
#
# print(cell.value)
# print(sheet.max_row)
#
# for row in range(1, sheet.max_row + 1):
#     print(row)
#
# for row in range(2, sheet.max_row + 1):
#     cell = sheet.cell(row, 3)
#     print(cell.value)
#
# for row in range(2, sheet.max_row + 1):
#     cell = sheet.cell(row, 3)
#     corrected_price = cell.value * 0.9
#     corrected_price_cell = sheet.cell(row, 4)
#     corrected_price_cell.value = corrected_price
#
# wb.save("transaction2.xlsx")

# ---------------------------------------------------------------

import openpyxl as xl
from openpyxl.chart import BarChart, Reference
# importing openpyxl library and give it a name as "xl"
# importing BarChart and Reference class from Chart package


def workbook_process(filename):
    wb = xl.load_workbook(filename)
    sheet = wb["Sheet1"]
    cell = sheet["a1"]
    # cell = sheet.cell(1, 1)  =  cell = wb["a1"]
    print(cell.value)
    print(sheet.max_row)
    # accessing excel sheet
    # load your excel workbook and create an object/instance of this workbook
    # accessing the first sheet of excel file and give it to a name
    # accessing the first cell of excel file and give it to a name

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        print(cell.value)
    # accessing individual cell that you want
    # checking how many rows in that sheet
    # we add a for loop that will generate number 2 to 4 using range function
    # we are accessing into the cells of column number 3 and create cell object
    # print cell value
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price
        # next we need to multiply each value with 0.9 and add it to a
        # ...newly created column here
        # we multiply cell value with 0.9 and initiate it to a variable
        # we need a reference for the cell of fourth column before create the
        # ...fourth column cell

    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, "F2")

    wb.save(filename)

    # create instances of Reference class and give 4 argument to specify the
    # ...column and row we want to select and then create an object of this instance
    # we select row from 2nd to end and select the column al 4 column
    # we create instance and object of BarChart class for use the class
    # we add the data into this BarChart object from Reference class we created
    # then we create the real chart into sheet with the "value" added into chart
    # ...and the "place" into sheet where we want the chart
    # we set the value for the new cell







